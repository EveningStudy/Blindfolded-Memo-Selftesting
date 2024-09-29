import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# input file
input_file = ''
df = pd.read_excel(input_file)


# output file
output_file = ''
wb = openpyxl.Workbook()
ws = wb.active

alphabets = 'ABCDEFGHIJKLMNOPQRSTWXYZ'
for i, letter in enumerate(alphabets):
    ws.cell(row=1, column=i + 2).value = letter
    ws.cell(row=i + 2, column=1).value = letter

skipped_codes = []
for index, row in df.iterrows():
    for col in range(0, len(row), 3):
        code = row[col]
        word = row[col + 1]
        pronunciation = row[col + 2]

        if pd.isna(code) or pd.isna(word) or pd.isna(pronunciation):
            continue

        if len(code) == 1:
            first_letter = code[0]
            if first_letter in alphabets:
                row_index = alphabets.index(first_letter) + 2
                col_index = alphabets.index(first_letter) + 2
                ws.cell(row=row_index, column=col_index).value = f"{word}\n{pronunciation}"
                ws.cell(row=row_index, column=col_index).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                               vertical='center')
            else:
                skipped_codes.append(code)
            continue

        first_letter = code[0]
        second_letter = code[1]

        if first_letter in alphabets and second_letter in alphabets:
            col_index = alphabets.index(first_letter) + 2
            row_index = alphabets.index(second_letter) + 2
            ws.cell(row=row_index, column=col_index).value = f"{word}\n{pronunciation}"
            ws.cell(row=row_index, column=col_index).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                           vertical='center')
        else:
            skipped_codes.append(code)


for i in range(2, len(alphabets) + 2):
    ws.column_dimensions[get_column_letter(i)].width = 20
    ws.row_dimensions[i].height = 40

wb.save(output_file)


